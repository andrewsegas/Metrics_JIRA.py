from jira_process import JiraProcess
import datetime
import numpy as np
import pandas as pd
import os
import config
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
from unidecode import unidecode

# ### Instância a classe de Auditoria de processo Protheus
jiraProcess = JiraProcess()

allCycle_Lead = []
allCycle_Lead.append(config.projectList[0])
allCycle_Lead.append(["Cycle Time"])  #1
allCycle_Lead.append(["Queue Time"])  #2
allCycle_Lead.append(["Lead Time"])  #3
allCycle_Lead.append(["Suporte Time"])  #4

for x in range(0,len(config.projectList[0])):
    if config.projectList[0][x] in 'mob|ba':  # get story with manutencao
        # Cycle e Fila
        cfilterCycle = '(' + config.projectList[1][x] + ') AND resolved >= startOfMonth(-1) AND resolved <= endOfMonth(-1) AND issuetype  in (Story, Manutenção, "Rejeição - Manutenção") AND cf[17100] is not EMPTY ORDER BY resolved ASC'

        cfilterLead = '(' + config.projectList[1][x] + ') AND resolved >= startOfMonth(-1) AND resolved <= endOfMonth(-1) AND issuetype in (Story, Manutenção, "Rejeição - Manutenção") AND cf[11043] is not EMPTY ORDER BY resolved ASC'
    else:
        #Cycle e Fila
        cfilterCycle = '(' + config.projectList[1][x] +') AND resolved >= startOfMonth(-1) AND resolved <= endOfMonth(-1) AND issuetype  in (Manutenção, "Rejeição - Manutenção") AND cf[17100] is not EMPTY ORDER BY resolved ASC'
        #Lead e Suporte
        cfilterLead = '(' + config.projectList[1][x] +') AND resolved >= startOfMonth(-1) AND resolved <= endOfMonth(-1) AND issuetype in (Manutenção, "Rejeição - Manutenção") AND cf[11043] is not EMPTY ORDER BY resolved ASC'

    # ### Obtém uma lista de issues conforme o filtro. ###
    print("Buscando issues no filtro para Cycle: " + config.projectList[0][x])
    issues = jiraProcess.getListIssues(cfilterCycle)
    print(f'Quantidade de Issues: {len(issues)}')

    wb = Workbook()
    sheet1 = wb.active
    sheet1.title = "Cycle"

    if(len(issues) > 0):
        issuesCycle = [jiraProcess.getCycle_Queue(x.key) for x in issues]

        d = []

        d.append(([x.key for x in issues]))  # 'Issue'
        d.append(([x[0] for x in issuesCycle]))  # 'Criação'
        d.append(([x[1] for x in issuesCycle]))  # 'Data Inicio Planejado'
        d.append(([x[2] for x in issuesCycle]))  # 'Resolvido'
        d.append(([x[3] for x in issuesCycle]))  # 'Cycle'
        d.append(([x[4] for x in issuesCycle]))  # 'Queue'

        sheet1.cell(1, 1).value = "Issues"
        sheet1.cell(1, 1).font = Font(bold=True)
        sheet1.cell(1, 2).value = "Created"
        sheet1.cell(1, 2).font = Font(bold=True)
        sheet1.cell(1, 3).value = "Data Inicio Planejado"
        sheet1.cell(1, 3).font = Font(bold=True)
        sheet1.cell(1, 4).value = "Resolved"
        sheet1.cell(1, 4).font = Font(bold=True)
        sheet1.cell(1, 5).value = "Cycle"
        sheet1.cell(1, 5).font = Font(bold=True)
        sheet1.cell(1, 6).value = "Queue"
        sheet1.cell(1, 6).font = Font(bold=True)

        for i, l in enumerate(d):
            for j, col in enumerate(l):
                if (i < 4):
                    sheet1.cell(j+2, i+1, col)
                elif (i == 4):
                    sheet1.cell(j+2, i+1, (f"=D{j+2}-C{j+2}"))
                else:
                    sheet1.cell(j+2, i+1, (f"=C{j+2}-B{j+2}"))

        allCycle_Lead[1].append((sum(d[4], datetime.timedelta(0,0))/len(d[4])))
        allCycle_Lead[2].append((sum(d[5], datetime.timedelta(0,0)) / len(d[5])))

        sheet1.cell(j+3, 5, allCycle_Lead[1][-1].days)
        sheet1.cell(j+3, 6, allCycle_Lead[2][-1].days)
    else:
        allCycle_Lead[1].append(datetime.timedelta(0))
        allCycle_Lead[2].append(datetime.timedelta(0))

    ### GET TransitiONS ###

    if config.lTransition:
        print("Buscando Transitions ")
        transition = [jiraProcess.getTransition(x.key) for x in issues]

        # Procura o status do projeto atual
        if config.projectList[0][x] in config.statusList[0]:
            ind = config.statusList[0].index(config.projectList[0][x])
        else:
            print("O Projeto " + config.projectList[0][x] + " nao tem Status para transition, sera utilizado o padrao")
            ind = config.statusList[0].index('padrao')

        if config.statusList[1][ind] != "0":
            # retira os espaços
            listStatus = [i.strip() for i in config.statusList[1][ind].split(',')]

            sheetT = wb.create_sheet("Transition")

            sheetT.cell(1, 1).value = "Issues"
            sheetT.cell(1, 1).font = Font(bold=True)

            for nT, sT in enumerate(listStatus):  # preenche cabecalho do status
                sheetT.cell(1, nT + 2).value = sT
                sheetT.cell(1, nT + 2).font = Font(bold=True)

            for nIssue, cont in enumerate(transition):  # Issue por issue
                sheetT.cell(nIssue + 2, 1, issues[nIssue].key)  # Preenche o codigo da issue
                for nStatus, sStatus in enumerate(listStatus):  # Status por Status
                    for item in cont:  # transition por transition
                        if sStatus.lower() in config.calculatorList[0]:
                            indCalc = config.calculatorList[0].index(sStatus.lower())
                            formula = config.calculatorList[1][indCalc].replace("$",str(nIssue + 2)) # da replace no $ para a linha
                            sheetT.cell(nIssue + 2, nStatus + 2,"=" + formula )  # Preenche a formula na tag [Calculator]
                            break
                        if unidecode(item[0].lower()) == unidecode(sStatus.lower()):
                            #print("achou")
                            dateT = datetime.datetime.strptime(item[1], '%Y-%m-%dT%H:%M:%S.%f%z')
                            sheetT.cell(nIssue + 2, nStatus + 2, dateT)  # Preenche a data
                            break
                    else:
                        print("Nao achou: " + unidecode(sStatus.lower()) + " Na issue " + issues[nIssue].key)
                        sheetT.cell(nIssue + 2, nStatus + 2, datetime.timedelta(0))  # Preenche o codigo da issue

    #Lead e Suporte

    # ### Obtém uma lista de issues conforme o filtro. ###
    print("Buscando issues no filtro para Lead: " + config.projectList[0][x])
    issues = jiraProcess.getListIssues(cfilterLead)
    print(f'Quantidade de Issues: {len(issues)}')

    sheet2 = wb.create_sheet("Lead")

    if (len(issues) > 0):
        issuesCycle = [jiraProcess.getLead_Suporte(x.key) for x in issues]

        ### Finished JIRA connection

        d = []

        d.append(([x.key for x in issues]))  # 'Issue'
        d.append(([x[0] for x in issuesCycle]))  # 'Criação'
        d.append(([x[1] for x in issuesCycle]))  # 'Data abertura ticket'
        d.append(([x[2] for x in issuesCycle]))  # 'Resolvido'
        d.append(([x[3] for x in issuesCycle]))  # 'Lead'
        d.append(([x[4] for x in issuesCycle]))  # 'Suporte'

        sheet2.cell(1, 1).value = "Issues"
        sheet2.cell(1, 1).font = Font(bold=True)
        sheet2.cell(1, 2).value = "Created"
        sheet2.cell(1, 2).font = Font(bold=True)
        sheet2.cell(1, 3).value = "Data Abertura Ticket"
        sheet2.cell(1, 3).font = Font(bold=True)
        sheet2.cell(1, 4).value = "Resolved"
        sheet2.cell(1, 4).font = Font(bold=True)
        sheet2.cell(1, 5).value = "Lead"
        sheet2.cell(1, 5).font = Font(bold=True)
        sheet2.cell(1, 6).value = "Suporte"
        sheet2.cell(1, 6).font = Font(bold=True)

        for i, l in enumerate(d):
            for j, col in enumerate(l):
                if (i < 4):
                    sheet2.cell(j+2, i+1, col)
                elif (i == 4):
                    sheet2.cell(j+2, i+1, (f"=D{j+2}-C{j+2}"))
                else:
                    sheet2.cell(j+2, i+1, (f"=B{j+2}-C{j+2}"))


        allCycle_Lead[3].append((sum(d[4], datetime.timedelta(0,0))/len(d[4])))
        allCycle_Lead[4].append((sum(d[5], datetime.timedelta(0,0)) / len(d[5])))

        sheet2.cell(j+3, 5, allCycle_Lead[3][-1].days)
        sheet2.cell(j+3, 6, allCycle_Lead[4][-1].days)
    else:
        allCycle_Lead[3].append(datetime.timedelta(0))
        allCycle_Lead[4].append(datetime.timedelta(0))

    wb.save(os.getcwd() + '/Result/metrics_'+ config.projectList[0][x] +'.xlsx')

wFinal = Workbook()
sheetTotal = wFinal.active
sheetTotal.title = "Total"

for i, l in enumerate(allCycle_Lead):
    for j, col in enumerate(l):
        if ( j>0 and i>0):
            sheetTotal.cell(j + 1, i + 1, col.days)
        else:
            sheetTotal.cell(j + 1, i + 1, col)
            sheetTotal.cell(j + 1, i + 1).font = Font(bold=True)

wFinal.save(os.getcwd() + '/Result/metrics_FINAL.xlsx')

jiraProcess.close()
